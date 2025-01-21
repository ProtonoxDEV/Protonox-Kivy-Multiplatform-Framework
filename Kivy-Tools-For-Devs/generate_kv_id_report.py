import os
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Directorio para almacenar reportes
REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")

# Archivos de entrada y salida
JSON_FILE = os.path.join(REPORTS_DIR, "ids.json")
EXCEL_FILE = os.path.join(REPORTS_DIR, "kv_id_report.xlsx")
HTML_FILE = os.path.join(REPORTS_DIR, "kv_id_report.html")


def load_from_json(input_file):
    """
    Carga los datos de un archivo JSON.
    """
    if os.path.exists(input_file):
        with open(input_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def generate_excel_report(id_map, output_file):
    """
    Genera un reporte Excel de los IDs encontrados y duplicados.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ID Report"
    ws.append(["ID", "Files", "Duplicated"])

    fill_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for id_, files in id_map.items():
        duplicated = len(files) > 1
        files_str = "\n".join(files)
        ws.append([id_, files_str, "Yes" if duplicated else "No"])
        if duplicated:
            for cell in ws[ws.max_row]:
                cell.fill = fill_amarillo

    wb.save(output_file)
    print(f"Excel report generated: {output_file}")


def generate_html_report(id_map, output_file):
    """
    Genera un reporte HTML de los IDs encontrados y duplicados.
    """
    with open(output_file, "w", encoding="utf-8") as f:
        f.write("<html><head><title>KV ID Report</title></head><body>")
        f.write("<h1>Kivy KV ID Report</h1>")

        if id_map:
            for id_, files in id_map.items():
                f.write(f"<h2>ID: {id_}</h2><ul>")
                for file in files:
                    f.write(f"<li>{file}</li>")
                f.write("</ul>")
        else:
            f.write("<p>No IDs found.</p>")

        f.write("</body></html>")

    print(f"HTML report generated: {output_file}")


def main():
    id_map = load_from_json(JSON_FILE)
    if not id_map:
        print("No data found in JSON. Please run `analyze_kv.py` first.")
        return

    generate_excel_report(id_map, EXCEL_FILE)
    generate_html_report(id_map, HTML_FILE)
    print("Reports generated successfully.")


if __name__ == "__main__":
    main()
